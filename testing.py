import face_recognition
import cv2
import numpy as np
import pickle
import datetime
from openpyxl import load_workbook

webcam_source = cv2.VideoCapture(0)

original_number_of_faces = 0
name_of_true_faces = ""
face_locations_new = []
amount_of_faces = 0
original_encodings = 0
unknown_face_encodings = []
# note, cv2 is BGR not RGB
priorities = {"High": (0, 0, 255),
              "Medium": (0, 165, 255),
              "Low": (0, 255, 255),
              }

black_listed_students = {}

black_book = load_workbook(filename="Attendance.xlsx")
blacklist = black_book["Blacklist"]
for i in blacklist.iter_rows(values_only=True):
    black_listed_students.update({i[0]: i[1]})
black_book.close()


def create_black_list():
    correct = False
    student_name = input("Please enter the name of the student: ")
    while not correct:
        student_priority = input("Please enter the level of Priority for this student, select either 'High', "
                                 "'Medium', or 'Low' "
                                 "for the priority level: ")
        if student_priority == 'High' or 'Medium' or 'Low':
            black_book_new = load_workbook(filename="Attendance.xlsx")
            new_black_list = black_book_new["Blacklist"]
            amount_of_rows = len(new_black_list["A"]) + 1
            new_black_list[f'A{amount_of_rows}'] = student_name
            new_black_list[f'B{amount_of_rows}'] = student_priority
            black_book_new.save(filename="Attendance.xlsx")
            correct = True
        else:
            print("The priority you entered was incorrect, "
                  "please check the capitalisation of your input and re-enter it")


def create_new_encoding():
    with open('grammar_database.dat', 'rb') as CreateNewEntryDatabase:
        grammar_face_encodings = pickle.load(CreateNewEntryDatabase)
    print("Note: Only .JPG file extensions should be used.")
    image_name = input("Please enter the image name: ")
    image_path = f'known_faces/class/{image_name}.JPG'
    new_image = face_recognition.load_image_file(image_path)
    print("Generating New User Information...")
    print("Please Wait...")
    grammar_face_encodings[image_name] = face_recognition.face_encodings(new_image, num_jitters=100)[0]
    with open("grammar_database.dat", "wb") as CreateNewEntryDatabase:
        pickle.dump(grammar_face_encodings, CreateNewEntryDatabase)
    print("New User Added")


def compare_faces():
    global name_of_true_faces, face_locations_new, unknown_face_encodings
    amount_continue = False
    while not amount_continue:
        ret, frame2 = webcam_source.read()
        rubber_small_frame = cv2.cvtColor(frame2, cv2.COLOR_BGR2RGB)
        face_locations_new = face_recognition.face_locations(rubber_small_frame, number_of_times_to_upsample=2)
        unknown_face_encodings = face_recognition.face_encodings(rubber_small_frame, face_locations_new)
        amount_encodings = len(unknown_face_encodings)
        if amount_encodings == amount_of_faces:
            amount_continue = True
        else:
            pass
    with open('grammar_database.dat', 'rb') as DataBaseOpened:
        grammar_face_encodings = pickle.load(DataBaseOpened)
    known_face_names = list(grammar_face_encodings.keys())
    known_face_encodings = np.array(list(grammar_face_encodings.values()))
    for faces in unknown_face_encodings:
        check_faces = face_recognition.compare_faces(known_face_encodings, faces)
        # any() searches through the list of check_faces to find if there is a True boolean.
        if any(check_faces):
            # .Index gets the Index of the True variables
            get_index = check_faces.index(True)
            name_of_true_faces = known_face_names[get_index]
            # locations_of_true_faces = face_locations_new[get_index]
            print(f'{name_of_true_faces} is in attendance')
            if name_of_true_faces in black_listed_students:
                print(
                    f'Student {name_of_true_faces} is a Priority student, '
                    f'with a {(black_listed_students[name_of_true_faces])} level of priority.')
            attendance_book = load_workbook(filename="Attendance.xlsx")
            class_name = attendance_book["SDD12"]
            date_time_tool = datetime.datetime.now().time()
            date_time_tool = str(date_time_tool)
            new_times = date_time_tool.split(":")
            hours = new_times[0]
            minutes = new_times[1]
            am_pm = "AM"
            if int(hours) > 12:
                am_pm = "PM"
                hours = int(hours) - 12
            count = 0
            for names in class_name.iter_rows(values_only=True):
                count += 1
                if name_of_true_faces == names[0]:
                    class_name[f'B{count}'] = "Present"
                    class_name[f'C{count}'] = f'{hours}:{minutes} {am_pm}'
            attendance_book.save(filename="Attendance.xlsx")
        else:
            print("An Unregistered user has been detected!\n"
                  "Please Adjust lighting and Re-Scan Face.")


def video_feed_display():
    global original_number_of_faces, original_encodings, amount_of_faces
    video_loop = True
    while video_loop is True:
        ret, frame = webcam_source.read()
        small_frame = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
        rgb_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame, number_of_times_to_upsample=2)
        amount_of_faces = len(face_locations)
        if amount_of_faces > original_number_of_faces:
            original_number_of_faces = amount_of_faces
            compare_faces()
        # Add a While loop until amount of face encodings match
        else:
            pass
        if name_of_true_faces in black_listed_students:
            priority_level = black_listed_students[name_of_true_faces]
            colour = priorities[priority_level]

        for (top, right, bottom, left) in face_locations:
            top *= 2
            right *= 2
            bottom *= 2
            left *= 2
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 128, 0), 2)
        cv2.imshow("God's Eye", frame)
        cv2.waitKey(1)

    webcam_source.release()
    cv2.destroyAllWindows()


print("Welcome to the God's Eye Program")
print("\n"
      "##############################################################\n"
      "#       _____           _ _           _____                  #\n"
      "#      |  __ \         | ( )         |  ___|                 #\n"
      "#      | |  \/ ___   __| |/ ___      | |__ _   _  ___        #\n"
      "#      | | __ / _ \ / _` | / __|     |  __| | | |/ _ \       #\n"
      "#      | |_\ \ (_) | (_| | \__ \     | |__| |_| |  __/       #\n"
      "#       \____/\___/ \__,_| |___/     \____/\__, |\___|       #\n"
      "#                                          __/ |            #\n"
      "#                                         |___/             #\n"
      "#############################################################\n")
print(
    "The God's Eye is a new form of Surveillance Technology, "
    "created for sole purpose of recording student attendance.")
user_continue = False
user_input = input("To Continue to a section of the code please select one of the following options:\n"
                   "a. Face Recognition\n"
                   "b. Create a new Face encoding\n"
                   "c. Create a new Black Listed student\n"
                   "d. Read Instructions on how to use program.\n")
if user_input == 'a':
    video_feed_display()
elif user_input == 'b':
    create_new_encoding()
elif user_input == 'c':
    create_black_list()
elif user_input == 'd':
    print("These are some instructions...")
else:
    print("Unfortunately the input entered did not match one of the available options."
          "Please select again.")
