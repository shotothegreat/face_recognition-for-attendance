import face_recognition
import cv2
import numpy
import pickle
import datetime
from openpyxl import load_workbook

face_locations = []
camera_index = 0
webcam_source = cv2.VideoCapture(camera_index)
students_present = []
original_number_of_faces = 0
name_of_true_faces = ""
face_locations_new = []
amount_of_faces = 0
original_encodings = 0
unknown_face_encodings = []
colour_of_face = []
set_tolerance = 0.6
colour = ()
# note, cv2 is BGR not RGB
priorities = {"High": (0, 0, 255),
              "Medium": (0, 165, 255),
              "Low": (0, 255, 255),
              }

time_table = {"RC": "08:35-08:48",
              "P1": "08:48-09:51",
              "P2": "09:51-10:54",
              "P3": "11:13-12:16",
              "P4": "12:16-13:19",
              "P5": "13:57-15:00",
              "P6": "17:57-24:01", }


# black_listed_students = {}
# This is the Black-List code -> will be implemented in version 2.0
# black_book = load_workbook(filename="Attendance.xlsx")
# blacklist = black_book["Blacklist"]
# for i in blacklist.iter_rows(values_only=True):
# black_listed_students.update({i[0]: i[1]})
# black_book.close()

def create_black_list():
    correct = False
    student_name = input("Please enter the name of the student: ")
    while not correct:
        correct = True
        student_priority = input("Please enter the level of Priority for this student, select either 'High', "
                                 "'Medium', or 'Low' "
                                 "for the priority level: ")
        if student_priority == 'High' or 'Medium' or 'Low':
            black_book_new = load_workbook(filename="Attendance.xlsx")
            new_black_list = black_book_new["Blacklist"]
            amount_of_rows = len(new_black_list["A"]) + 1
            new_black_list[f'A{amount_of_rows}'] = student_name
            new_black_list[f'B{amount_of_rows}'] = student_priority
            black_book_new.save(filename="Attendance.xlsx")
        else:
            print("The priority you entered was incorrect, "
                  "please check the capitalisation of your input and re-enter it")
            # Needs to keep scanning until a face is found


with open('grammar_database.dat', 'rb') as DataBaseOpened:
    grammar_face_encodings = pickle.load(DataBaseOpened)
    known_face_names = list(grammar_face_encodings.keys())
    known_face_encodings = numpy.array(list(grammar_face_encodings.values()))


def create_new_encoding():
    print("Note: Only .JPG file extensions should be used.")
    image_name = input("Please enter the image name: ")
    image_path = f'known_faces/class/{image_name}.JPG'
    new_image = face_recognition.load_image_file(image_path)
    print("Generating New User Information (Approximately 1-3 minutes)...")
    print("Please Wait...")
    grammar_face_encodings[image_name] = face_recognition.face_encodings(new_image, num_jitters=100)[0]
    with open("grammar_database.dat", "wb") as CreateNewEntryDatabase:
        pickle.dump(grammar_face_encodings, CreateNewEntryDatabase)
    print("New User Added")


def backend_faster_recognition():
    original_number_of_faces_2 = 0
    face_recognition_loop = True
    while face_recognition_loop is True:
        ret, frame = webcam_source.read()
        small_frame = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
        rgb_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        face_locations_2 = face_recognition.face_locations(rgb_frame, number_of_times_to_upsample=2)
        amount_of_faces_2 = len(face_locations_2)
        if amount_of_faces_2 != original_number_of_faces_2 and amount_of_faces_2 > 0:
            original_number_of_faces_2 = amount_of_faces_2
            unknown_face_encodings_2 = face_recognition.face_encodings(rgb_frame, face_locations_2)
            for faces in unknown_face_encodings_2:
                check_faces = face_recognition.compare_faces(known_face_encodings, faces)
                if any(check_faces):
                    get_index = check_faces.index(True)
                    name_of_true_faces_2 = known_face_names[get_index]
                    attendance_book = load_workbook(filename="Attendance.xlsx")
                    class_name = attendance_book[new_student_class]
                    date_time_tool = str(datetime.datetime.now().time())
                    date_time_tool = str(date_time_tool)
                    new_times = date_time_tool.split(":")
                    hours = new_times[0]
                    minutes = new_times[1]
                    am_pm = "AM"
                    if int(hours) > 12:
                        am_pm = "PM"
                        hours = int(hours) - 12
                    count = 0
                    new_check_time = str(datetime.datetime.now().time())
                    new_check_time = new_check_time.split(":")
                    for new_times in time_table:
                        new_roll = time_table[new_times].split("-")
                        new_start_time = ((int(new_roll[0][:2])) * 60) + (int(new_roll[0][-2:]))
                        new_end_time = ((int(new_roll[1][:2])) * 60) + (int(new_roll[1][-2:]))
                        new_current_time = ((int(new_check_time[0]) * 60) + (int(new_check_time[1])))
                        if (int(new_start_time)) <= (int(new_current_time)) <= (int(new_end_time)):
                            for names in class_name.iter_rows(values_only=True):
                                count += 1
                                if name_of_true_faces_2 == names[0]:
                                    presence = (class_name[f'B{count}']).value
                                    if presence == "Absent":
                                        print(f'{name_of_true_faces_2} is present.')
                                        class_name[f'B{count}'] = f'Present in {new_times}'
                                        class_name[f'C{count}'] = f'{hours}:{minutes} {am_pm}'
                                        attendance_book.save(filename="Attendance.xlsx")
                                        break


def compare_faces():
    global name_of_true_faces, face_locations_new, unknown_face_encodings, amount_of_faces, colour_of_face, colour
    amount_continue = False
    while not amount_continue:
        ret, frame2 = webcam_source.read()
        rubber_small_frame = cv2.cvtColor(frame2, cv2.COLOR_BGR2RGB)
        face_locations_new = face_recognition.face_locations(rubber_small_frame, number_of_times_to_upsample=2)
        unknown_face_encodings = face_recognition.face_encodings(rubber_small_frame, face_locations_new)
        amount_encodings = len(unknown_face_encodings)
        if amount_encodings == amount_of_faces:
            amount_continue = True
    colour = (0, 0, 128)
    for faces in unknown_face_encodings:
        check_faces = face_recognition.compare_faces(known_face_encodings, faces, tolerance=set_tolerance)
        # any() searches through the list of check_faces to find if there is a True boolean.
        if any(check_faces):
            # .Index gets the Index of the True variables
            get_index = check_faces.index(True)
            name_of_true_faces = known_face_names[get_index]
            colour = (0, 128, 0)
            # if name_of_true_faces in black_listed_students:
            # print(
            # f'Student {name_of_true_faces} is a Priority student,'
            # f' with a {(black_listed_students[name_of_true_faces])} level of priority.')
            attendance_book = load_workbook(filename="Attendance.xlsx")
            class_name = attendance_book[new_student_class]
            date_time_tool = str(datetime.datetime.now().time())
            date_time_tool = str(date_time_tool)
            new_times = date_time_tool.split(":")
            hours = new_times[0]
            minutes = new_times[1]
            am_pm = "AM"
            if int(hours) > 12:
                am_pm = "PM"
                hours = int(hours) - 12
            count = 0
            new_check_time = str(datetime.datetime.now().time())
            new_check_time = new_check_time.split(":")
            for new_times in time_table:
                new_roll = time_table[new_times].split("-")
                new_start_time = ((int(new_roll[0][:2])) * 60) + (int(new_roll[0][-2:]))
                new_end_time = ((int(new_roll[1][:2])) * 60) + (int(new_roll[1][-2:]))
                new_current_time = ((int(new_check_time[0]) * 60) + (int(new_check_time[1])))
                if (int(new_start_time)) <= (int(new_current_time)) <= (int(new_end_time)):
                    for names in class_name.iter_rows(values_only=True):
                        count += 1
                        if name_of_true_faces == names[0]:
                            presence = (class_name[f'B{count}']).value
                            if presence == "Absent":
                                print(f'{name_of_true_faces} is in attendance')
                                class_name[f'B{count}'] = f'Present in {new_times}'
                                class_name[f'C{count}'] = f'{hours}:{minutes} {am_pm}'
                                attendance_book.save(filename="Attendance.xlsx")
                                break
        else:
            print("An Unregistered user has been detected!\n"
                  "Please Adjust lighting and Re-Scan Face.")
            amount_of_faces -= 1
        colour_of_face.append(colour)


def video_feed_display():
    global original_number_of_faces, original_encodings, amount_of_faces, face_locations, colour
    video_loop = True
    while video_loop is True:
        ret, frame = webcam_source.read()
        small_frame = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
        rgb_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame, number_of_times_to_upsample=2)
        amount_of_faces = len(face_locations)
        if amount_of_faces != original_number_of_faces and amount_of_faces > 0:
            original_number_of_faces = amount_of_faces
            compare_faces()
        # Add a While loop until amount of face encodings match
        else:
            pass
        for (x, y, w, h), colour in zip(face_locations, colour_of_face):
            x *= 2
            y *= 2
            w *= 2
            h *= 2
            cv2.rectangle(frame, (h, x), (y, w), colour, 2)
        cv2.imshow("God's Eye", frame)
        cv2.waitKey(1)
    webcam_source.release()
    cv2.destroyAllWindows()


print("Welcome to the God's Eye Program")
print("\n"
      "                                                         ",
      "        _____           _ _           _____              ",
      "       |  __ \         | ( )         |  ___|             ",
      "       | |  \/ ___   __| |/ ___      | |__ _   _  ___    ",
      "       | | __ / _ \ / _` | / __|     |  __| | | |/ _ \   ",
      "       | |_\ \ (_) | (_| | \__ \     | |__| |_| |  __/   ",
      "        \____/\___/ \__,_| |___/     \____/\__, |\___|   ",
      "                                           __/ |         ",
      "                                          |___/          ",
      "                                                         ", sep='\n')
print(
    "The God's Eye is a new form of Surveillance Technology, "
    "created for sole purpose of recording student attendance.")
class_correct = False
while not class_correct:
    student_class = input("Please Enter Class e.g.'12SDD1: ").upper()
    check_excel_exists = load_workbook(filename="Attendance.xlsx")
    for number_of_sheets in check_excel_exists:
        try:
            var = check_excel_exists[student_class]
            class_correct = True
        except KeyError:
            print("The Class you entered doesn't exist. Please try again.")
    check_excel_exists.save(filename="Attendance.xlsx")

user_continue = False
while user_continue is False:
    user_input = input("To Continue to a section of the code please select one of the following options:\n"
                       "a. Face Recognition\n"
                       "b. Create a new Face encoding\n"
                       "c. Create a new Black Listed student\n"
                       "d. Settings.\n"
                       "e. Quit Program\n").lower()
    if user_input == 'a':
        with open("Settings.txt", "r+") as settings_file:
            lines = settings_file.readlines()
            for line in lines:
                new_line = line.split("|")
                if new_line[0] == "webcam":
                    camera_index = (int(new_line[1]))
                elif new_line[0] == "tolerance":
                    set_tolerance = float((new_line[1]))
        check_time = str(datetime.datetime.now().time())
        check_time = check_time.split(":")
        for times in time_table:
            roll = time_table[times].split("-")
            start_time = ((int(roll[0][:2])) * 60) + (int(roll[0][-2:]))
            end_time = ((int(roll[1][:2])) * 60) + (int(roll[1][-2:]))
            current_time = ((int(check_time[0]) * 60) + (int(check_time[1])))
            if (int(start_time)) <= (int(current_time)) <= (int(end_time)):
                print(f"It is currently: {times}")
                break
        else:
            print("It is not currently class time. If this is wrong, "
                  "please change the class times within settings.")
        new_copy = load_workbook(filename="Attendance.xlsx")
        sheet_to_copy = new_copy[student_class]
        sheet_to_change_name = new_copy.copy_worksheet(sheet_to_copy)
        check_date = str(datetime.datetime.now().date())
        check_date = check_date.split("-")
        new_student_class = f'{student_class}_{check_date[2]}.{check_date[1]}.{check_date[0]}'
        sheet_to_change_name.title = new_student_class
        new_copy.save(filename="Attendance.xlsx")
        user_choice = False
        while user_choice is False:
            which_option_of_recognition = input("Press [A] for live video feed, "
                                                "or press [B] for Non-Visual-Display version ").lower()
            if which_option_of_recognition == 'a':
                video_feed_display()
            elif which_option_of_recognition == 'b':
                backend_faster_recognition()
            else:
                print("The button you entered did not match one of the options, please try again.")
    elif user_input == 'b':
        create_new_encoding()
    elif user_input == 'c':
        print("Unfortunately, this option is not available in the current version of God's Eye.")
        # create_black_list()
    elif user_input == 'd':
        print(f"""Current Settings are:\n
        Webcam: System_Default_Camera: {camera_index} \n
        Tolerance: {set_tolerance} \n
        Timetable: {time_table} \n""")
        correct_setting = False
        setting = input("Which setting would you like to change? or press [Q] to cancel. ").lower()
        if setting == 'webcam':
            change_cam = input("Press [C] to change the Camera, or [R] to reset the camera. ").lower()
            if change_cam == 'c':
                print("Changing the Camera to the next available camera...")
                camera_index += 1
            elif change_cam == 'r':
                print("Resetting Camera index to System Default...")
                camera_index = 0
            else:
                print("Invalid input, please try again later.")
            with open("Settings.txt", "r+") as settings_file_2:
                lines_2 = settings_file_2.readlines()
                for line_2 in lines_2:
                    new_line_2 = line_2.split("|")
                    if new_line_2[0] == "tolerance":
                        var2_2 = line_2
                var3_2 = f'webcam|{camera_index}'
                new_file_information_2 = "\n" + var2_2 + "\n" + var3_2
                settings_file_2.truncate(0)
                settings_file_2.write(new_file_information_2)
                print(f"Camera has been successfully set to System_Default No.{camera_index}")
        elif setting == 'tolerance':
            change_tolerance_check = False
            while change_tolerance_check is False:
                change_tolerance = input("Press [R] to reset tolerance, or press [C] to change tolerance. ").lower()
                if change_tolerance == 'r':
                    set_tolerance = 0.6
                elif change_tolerance == 'c':
                    print("Suggested tolerances should be in a range of 0.4 - 0.7")
                    print("Before setting a full decimal change to tolerance e.g. 0.6 to 0.7,"
                          " try changing the hundredth of the tolerance e.g. 0.6 to 0.65")
                    new_tolerance = input("Please enter a new tolerance. ")
                    new_tolerance = float(new_tolerance)
                    if new_tolerance >= 1:
                        print("The value entered was invalid as value should be in range of 0.4 - 0.7")
                        print("Try again.")
                    else:
                        set_tolerance = new_tolerance
                with open("Settings.txt", "r+") as settings_file_1:
                    lines_1 = settings_file_1.readlines()
                    for line_1 in lines_1:
                        new_line_1 = line_1.split("|")
                        if new_line_1[0] == "webcam":
                            var1 = line_1
                    var3 = f'tolerance|{set_tolerance}'
                    new_file_information = "\n" + var1 + "\n" + var3
                    settings_file_1.truncate(0)
                    settings_file_1.write(new_file_information)
                    print(f"Tolerance has been successfully set to: {set_tolerance}")
                    break
        elif setting == 'timetable':
            print("Unfortunately in this version of the code, the timetable is not editable.")
            # timetable_continue = False
            # while not timetable_continue:
            # change_timetable = input("Enter [C] to change timetable settings or [Q] to cancel. ").lower()
            # if change_timetable == 'c':
            # new_change_timetable_continue = False
            # while not new_change_timetable_continue:
            # new_change_timetable = input("Enter [T] to change timetable times, "
            # "or press [N] to create a new entry. ").lower()
            # if new_change_timetable == 't':
            # print("The current periods and respective times are:")
            # for entries in time_table:
            #  print(f'{entries}: {time_table[entries]}')
            # which_change = input("Which Period time would you like to change? ").upper()
            # if which_change in time_table:
            #  time_changer_1 = input("Please enter the time at which the period starts e.g. 08:43: ")
            # time_changer_2 = input("Please enter the time at which the period ends e.g. 09:20: ")
            # time_table[which_change] = f'{time_changer_1}-{time_changer_2}'
            # break
            # else:
            # print("That period is not currently in the timetable. Please re-enter.")
            # elif new_change_timetable == 'n':
            # print("Placeholder")
            # else:
            # print("The input you entered was incorrect, please retry.")
            # elif change_timetable == 'q':
            #   break
            #  else:
            # print("The input you entered was incorrect, please retry.")
        elif setting == 'q':
            correct_setting = True
        else:
            print("The value you entered did not match our settings, "
                  "please select either: timetable, webcam or tolerance. or press [Q] to go back to the main screen.")
    elif user_input == 'e':
        exit()
    else:
        print("Unfortunately the input entered did not match one of the available options. "
              "Please select again.")
